from django.shortcuts import render
from django.db.models.aggregates import Max

from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status, generics
from rest_framework.decorators import permission_classes
from rest_framework.permissions import IsAuthenticated, IsAdminUser

from users.models import User
from .models import Attendance, Logs
import datetime, time

from io import StringIO
from django.http import HttpResponse
from xlsxwriter.workbook import Workbook
import openpyxl

# Create your views here.

class login(APIView):
    """
    User day-in login
    """
    permission_classes = [IsAuthenticated]

    def post(self, request, format=None):
        try:
            attendance = Attendance.objects.get(user__pk = request.user.id, date = datetime.date.today())
        except:
            user = User.objects.get(username = request.user.username)
            start_minutes = user.start_time.hour*60 + user.start_time.minute
            end_minutes = user.end_time.hour*60 + user.end_time.minute
            ctime = (time.gmtime().tm_hour+5)*60 + time.gmtime().tm_min + 30
            if ctime >= start_minutes:
                attendance = Attendance(user = user, date = datetime.date.today(), 
                attendance_status = 'WORKING', working_window = end_minutes - start_minutes)
                attendance.save()
                attendance = Attendance.objects.get(user__pk = request.user.id, date = datetime.date.today())
                log = Logs(logid = attendance, type = 1)
                log.save()
                resp = {'message': 'Hello, ' + request.user.first_name}
                return Response(resp)
            else:
                resp = {'message': 'Please login on or after ' + user.start_time}
                return Response(resp)
        else:
            resp = {'message': 'You are already logged in.'}
            return Response(resp)

class logout(APIView):
    """
    User day-out login
    """
    permission_classes = [IsAuthenticated]

    def post(self, request, format=None):
        try:
            attendance = Attendance.objects.get(user__pk = request.user.id, date = datetime.date.today())
        except:
            resp = {'message': 'Please login before you logout'}
            return Response(resp)
        attendance = Attendance.objects.get(user__pk = request.user.id, date = datetime.date.today())
        user = User.objects.get(username = request.user.username)
        start_minutes = user.start_time.hour*60 + user.start_time.minute
        end_minutes = user.end_time.hour*60 + user.end_time.minute
        ctime = (time.gmtime().tm_hour+5)*60 + time.gmtime().tm_min + 30
        if attendance.is_completed == True:
                resp = {'message': 'Logged out for the day.'}
                return Response(resp)
        if ctime >= end_minutes:
            minutes = 0
            lastactive = Logs.objects.filter(logid__user__pk = request.user.id).order_by('id').last()
            ctime = (time.gmtime().tm_hour+5)*60 + time.gmtime().tm_min + 30
            ptime = lastactive.time.hour*60 + lastactive.time.minute
            minutes = ctime - ptime
            minutes += int(lastactive.active_time)
            attendance.total_logouts += 1
            attendance.active_window = minutes
            start_minutes = user.start_time.hour*60 + user.start_time.minute
            end_minutes = user.end_time.hour*60 + user.end_time.minute
            if minutes > end_minutes - start_minutes:
                attendance.attendance_status = 'PRESENT'
            else:
                attendance.attendance_status = 'ABSENT' 
            attendance.is_completed = True    
            attendance.save()       
            log = Logs(logid = attendance, type = 0, active_time = minutes)
            log.save()
            resp = {'message': 'Have a nice time, ' + request.user.first_name}
            return Response(resp)
        else:
            resp = {'message': 'Please logout on or after ' + str(user.end_time)}
            return Response(resp)
        

class iLogin(APIView):
    """
    Intermediate login
    """
    permission_classes = [IsAuthenticated]

    def post(self, request, format=None):
        try:
            attendance = Attendance.objects.get(user__pk = request.user.id, date = datetime.date.today())
        except:
            resp = {'message': 'Your workday has not started. Please login.'}
            return Response(resp)
        else:
            # user = User.objects.get(username = request.user.username)
            attendance = Attendance.objects.get(user__pk = request.user.id, date = datetime.date.today())
            if attendance.is_completed == True:
                resp = {'message': 'Logged out for the day.'}
                return Response(resp)
            logs = Logs.objects.filter(logid__user__pk = request.user.id).order_by('id').last()
            if logs.type == 0:
                log = Logs(logid = attendance, type = 1)
                log.save()
                resp = {'message': 'Welcome back, ' + request.user.first_name}
                attendance.attendance_status = 'WORKING'
                attendance.total_logins += 1
                attendance.save()
                return Response(resp)
            else:
                resp = {'message': 'You are already logged in, ' + request.user.first_name}
                return Response(resp)

class iLogout(APIView):
    """
        Intermediate logout
    """
    permission_classes = [IsAuthenticated]

    def post(self, request, format=None):
        try:
            attendance = Attendance.objects.get(user__pk = request.user.id, date = datetime.date.today())
        except:
            resp = {'message': 'Your workday has not started. Please login.'}
            return Response(resp)
        else:
            # user = User.objects.get(username = request.user.username)
            attendance = Attendance.objects.get(user__pk = request.user.id, date = datetime.date.today())
            if attendance.is_completed == True:
                resp = {'message': 'Logged out for the day.'}
                return Response(resp)
            logs = Logs.objects.filter(logid__user__pk = request.user.id).order_by('id').last()
            if logs.type == 1:
                minutes = 0
                ctime = (time.gmtime().tm_hour+5)*60 + time.gmtime().tm_min + 30
                ptime = logs.time.hour*60 + logs.time.minute
                minutes = ctime - ptime
                minutes += int(logs.active_time)
                attendance.total_logouts += 1
                attendance.active_window = minutes
                attendance.attendance_status = 'BREAK'
                attendance.save()
                log = Logs(logid = attendance, type = 0)
                log.save()
                resp = {'message': 'Take a break, ' + request.user.first_name}
                return Response(resp)
            else:
                resp = {'message': 'You are already in break-time, ' + request.user.first_name}
                return Response(resp)

class userLogs(generics.ListAPIView):

    permission_classes = [IsAuthenticated]

    def get(self, request, format=None):
        """
        This view should return a list of all the logs
        for the currently authenticated user.
        """
        date = self.request.GET['date']
        if date is None:
            date = datetime.date.today()
        else:
            date = datetime.datetime.strptime(date, '%d-%m-%Y')
        user = self.request.user.id
        try:
            attendance = Attendance.objects.get(user__pk = request.user.id, date = date)
            userDetails = {'User ID': attendance.user.id, 'No of logins':attendance.total_logins,
                            'No of logouts':attendance.total_logouts, 
                            'Final status':attendance.attendance_status,
                            'Fixed start time':attendance.user.start_time,
                            'Fixed end time':attendance.user.end_time,
                            'Working window': attendance.working_window/60,
                            'Cummulative login session': attendance.active_window/60
                            }

            login = Logs.objects.filter(logid__user__pk=user, logid__date = date, type = True).values_list('time')
            logout = Logs.objects.filter(logid__user__pk=user, logid__date = date, type = False).values_list('time')
            login_list = []
            for x in range(len(login)):
                login_list.append(login[x][0])
            logout_list = []
            for x in range(len(logout)):
                logout_list.append(logout[x][0])
            log_list = {'user details': userDetails, 'login list': login_list, 'logout list':logout_list}

            return Response(log_list)
        except:
            resp = {'message': f"No data for {date.strftime('%d-%m-%Y')} found."}
            return Response(resp)

class allLogs(APIView):
    """
        All user logs
    """
    permission_classes = [IsAdminUser]

    def get(self, request, format=None):
        """
        This view should return a list of all the logs
        for the currently authenticated user.
        """
        date = self.request.GET['date']
        if date is None:
            date = datetime.date.today()
        else:
            date = datetime.datetime.strptime(date, '%d-%m-%Y')
        try:
            all_attendance = Attendance.objects.filter(date = date)
            log_list = {}
            for attendance in all_attendance:
                userDetails = {'User ID': attendance.user.id, 'No of logins':attendance.total_logins,
                                'No of logouts':attendance.total_logouts, 
                                'Final status':attendance.attendance_status,
                                'Fixed start time':attendance.user.start_time,
                                'Fixed end time':attendance.user.end_time,
                                'Working window': attendance.working_window/60,
                                'Cummulative login session': attendance.active_window/60
                                }

                login = Logs.objects.filter(logid__user__pk=attendance.user.id, logid__date = date, type = True).values_list('time')
                logout = Logs.objects.filter(logid__user__pk=attendance.user.id, logid__date = date, type = False).values_list('time')
                login_list = []
                for x in range(len(login)):
                    login_list.append(login[x][0])
                logout_list = []
                for x in range(len(logout)):
                    logout_list.append(logout[x][0])
                # log_list = {attendance.user.id : {'user details': userDetails, 'login list': login_list, 'logout list':logout_list}}
                log_list.update({attendance.user.id : {'user details': userDetails, 'login list': login_list, 'logout list':logout_list}})
            return Response(log_list)
        except:
            resp = {'message': f"No data for {date.strftime('%d-%m-%Y')} found."}
            return Response(resp)

class logDelete(APIView):
    """
    Clear Attendance and Logs tables
    """
    # permission_classes = [IsAuthenticated]

    def post(self, request, format=None):
        Attendance.objects.all().delete()
        Logs.objects.all().delete()
        resp = {'message': 'Flushed Logs and Attendance records.'}
        return Response(resp)

class changeWW(APIView):
    """
    View to change start_time and end_time
    """
    permission_classes = [IsAuthenticated]
    
    def post(self, request, format=None):    
        try:
            start = datetime.datetime.strptime(self.request.GET['start'], '%H:%M:%S')
            end = datetime.datetime.strptime(self.request.GET['end'], '%H:%M:%S')
        except:
            resp = {'message': 'start_time and end_time parameters not found.'}
            return Response(resp)
        else:
            user = User.objects.get(pk = request.user.id)
            user.start_time = start
            user.end_time = end
            user.save()
            start_minutes = start.hour*60 + start.minute
            end_minutes = end.hour*60 + end.minute
            attendance = Attendance.objects.get(user__pk = request.user.id, date = datetime.date.today())
            attendance.working_window = end_minutes - start_minutes
            attendance.save()
            resp = {'message': 'Time window updated.'}
            return Response(resp)

class downloadReport(APIView):
    """
    View to download login session report in excel format
    """
    permission_classes = [IsAdminUser]
    
    def get(self, request, format=None):    
        date = self.request.GET['date']
        if date is None:
            date = datetime.date.today()
        else:
            date = datetime.datetime.strptime(date, '%d-%m-%Y')
        try:
            all_attendance = Attendance.objects.filter(date = date)
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = f"attachment; filename={self.request.GET['date']}.xlsx"
            book = Workbook(response, {'in_memory': True})
            sheet = book.add_worksheet('Logs') 
        except:
            resp = {'message': f"No data for {date.strftime('%d-%m-%Y')} found."}
            return Response(resp)
        else:
            headlist = ['User ID', 'No of logins', 'No of logouts', 'Final status', 'Fixed start time', 'Fixed end time', 'Working window', 'Cummulative login session']
            for x in range(len(headlist)):
                sheet.write(0, x, headlist[x])

            row = 1
            for attendance in all_attendance:
                userDetails = {'User ID': attendance.user.id, 'No of logins':attendance.total_logins,
                                'No of logouts':attendance.total_logouts, 
                                'Final status':attendance.attendance_status,
                                'Fixed start time':attendance.user.start_time,
                                'Fixed end time':attendance.user.end_time,
                                'Working window': attendance.working_window/60,
                                'Cummulative login session': attendance.active_window/60
                                }
                col = 0
                for keys in  userDetails:
                    sheet.write(row, col, userDetails[keys])
                    col += 1
                row += 1
            book.close()
            return response

class uploadWW(APIView):
    """
    View to bulk upload user login and logout times
    """
    permission_classes = [IsAdminUser]
    
    def get(self, request, format=None):  
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = "attachment; filename=upload.xlsx"
        book = Workbook(response, {'in_memory': True})
        sheet = book.add_worksheet('Logs') 

        headlist = ['User ID', 'Start time', 'End time']
        for x in range(len(headlist)):
            sheet.write(0, x, headlist[x])
        book.close()
        return response 

    def post(self, request, format=None): 
        
        excel_file = request.FILES["excel_file"]

        # you may put validations here to check extension or file size

        wb = openpyxl.load_workbook(excel_file)

        # getting a particular sheet by name out of many sheets
        worksheet = wb["Logs"]
        # print(worksheet)

        excel_data = list()
        # iterating over the rows and
        # getting value from each cell in row
        cnt = 0
        message = {}
        valid = True
        for row in worksheet.iter_rows():
            row_data = list()
            if cnt == 0:
                row_cnt = 0
                for cell in row:
                    if str(cell.value) != 'User ID' and row_cnt == 0:
                        valid = False
                    elif str(cell.value) != 'Start time' and row_cnt == 1:
                        valid = False
                    elif str(cell.value) != 'End time' and row_cnt == 2:
                        valid = False
                    row_cnt += 1
            elif cnt > 0 and valid == True:
                try:
                    user = User.objects.get(pk = row[0].value)
                except:
                    message.update({f'UserID {row[0].value}':  'User not available'})
                else:
                    try:
                        attendance = Attendance.objects.get(user__pk = row[0].value, date = datetime.date.today())
                    except:
                        pass
                    row_cnt = 0
                    for cell in row:
                        row_data.append(str(cell.value))
                        if row_cnt == 1:
                            start = datetime.datetime.strptime(str(cell.value), '%H:%M:%S')
                            user.start_time = start
                            start_minutes = start.hour*60 + start.minute
                        elif row_cnt == 2:
                            end = datetime.datetime.strptime(str(cell.value), '%H:%M:%S')
                            user.end_time = end
                            end_minutes = end.hour*60 + end.minute
                            user.save()
                            attendance.working_window = end_minutes - start_minutes
                            attendance.save()  
                            message.update({f'User{row[0].value}': 'Time window updated.'}) 
                        row_cnt += 1                                                      
            else:
                resp = {'message': 'Provide data in right'}
                return Response(resp)
            excel_data.append(row_data)
            cnt += 1
        print(excel_data)
        message.update({'message': 'Upload successful'})
        return Response(message)